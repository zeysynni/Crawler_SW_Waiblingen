"""Convert the colleagues' knowledge-base spreadsheet to markdown.

    uv run python Excels/xlsx2md.py

This **runs in CI before every weekly crawl** (`.gitlab-ci.yml`), so replacing
`Knowledge Base.xlsx` in this folder is all it takes to refresh the knowledge
base — no local run, no hand-edited markdown. Because it is automated, previous
output is deleted before regenerating (`_clean_generated`): a renamed group
produces a new filename, and the superseded file would otherwise linger in
`static/` and keep being uploaded for ever.

The sheet ('Zusätzliche Informationen / Wissensdatenbank für den FAQ-Bot') is a
two-column list — 'Thema / Kategorie' and 'Inhalt / Wissen' — written by hand by
colleagues specifically as extra material for the FAQ bot. Each topic is named
'<Gruppe> – <Unterthema>', so the rows group into subject files.

One file per group, not one file per row and not one big file: the uploader makes
each file exactly one retrieval chunk, so a single 6.5k file would put all 18
unrelated topics into one chunk (a Freibad question would match a chunk that is
mostly Einspeiseanlagen). Per-group files match the granularity of the crawled
pages, where one page/chunk likewise covers several related sub-topics.

Writes into `static/`, from where the normal pipeline picks it up:
`main.py:copy_static()` copies `static/*.md` into `outputs/clean/` and `--upload`
pushes them like any crawled page. (Writing straight to `outputs/clean/` would
*not* work — it is gitignored and rebuilt each run, and `main.py` builds the
upload list from crawled pages + `static/*.md` only.)

`openpyxl` is a locked dependency in the `convert` group (`uv sync --group
convert`) — pinned rather than pulled per-invocation, so a library update cannot
silently change what a weekly automated run writes into a customer-facing
knowledge base. The workbook is opened with `data_only=True` so a formula cell yields
its computed value rather than '=SUM(...)'; note Excel only caches those values
on save, so a formula edited and never saved would read as empty (this sheet has
no formulas — the check below would catch it).

Output shape follows the crawled pages (see `clean.clean_markdown`): the h1 is
the entry's place in the hierarchy, sub-topics are `##`, no hyperlinks.
"""

import logging
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from clean import slug, strip_links  # noqa: E402

log = logging.getLogger("crawler")

XLSX_DIR = Path(__file__).resolve().parent
STATIC_DIR = XLSX_DIR.parent / "static"
WORKBOOK = XLSX_DIR / "Knowledge Base.xlsx"

# Root of the h1 hierarchy. These rows are the colleagues' supplementary
# knowledge, not crawled website content, so they carry their own root rather
# than blending into 'Privatkunden'/'Netze'.
HIERARCHY = "Wissensdatenbank"

# The two columns we need, matched on the header row so a reordered or renamed
# column fails loudly instead of silently producing empty files.
HEADER_TOPIC = "Thema / Kategorie"
HEADER_CONTENT = "Inhalt / Wissen"

MAX_CHUNK = 8192   # uploader.MAX_CHUNK — above this the KB API splits the file

# '<Gruppe> – <Unterthema>'. The sheet uses an en dash; a plain hyphen is
# accepted too, but only when spaced, so compounds ('Gäste-WLAN') stay intact.
_TOPIC_SPLIT = re.compile(r"\s+[–—-]\s+")

_UMLAUTS = str.maketrans({"ä": "ae", "ö": "oe", "ü": "ue", "Ä": "Ae",
                          "Ö": "Oe", "Ü": "Ue", "ß": "ss"})


def ascii_name(text: str) -> str:
    """Filename chunk, ASCII only: 'Elektromobilität' -> Elektromobilitaet.

    Umlauts are transliterated the German way so the filename cannot depend on
    Unicode normalisation form (macOS hands back decomposed names, Linux stores
    whatever bytes it is given). The knowledge base is keyed by filename and the
    uploader deletes/replaces by it, so a name that composes differently on the
    CI runner than locally would orphan the old file and upload a duplicate.
    Deliberately duplicated from `PDFs/pdf2md.py` to keep each one-off
    conversion folder self-contained. The h1 inside the file keeps real umlauts.
    """
    folded = unicodedata.normalize("NFC", text).translate(_UMLAUTS)
    stripped = "".join(c for c in unicodedata.normalize("NFD", folded)
                       if not unicodedata.combining(c))
    return slug(stripped.encode("ascii", "ignore").decode())


def _cell_text(value) -> str:
    """Normalise one cell to markdown-ready text.

    Collapses the runs of spaces that hand-typing leaves behind, but keeps
    in-cell line breaks as paragraph breaks (one row uses them to separate the
    explanation from the 'send us an e-mail at …' instruction).
    """
    if value is None:
        return ""
    paragraphs = [re.sub(r"[^\S\n]+", " ", p).strip()
                  for p in str(value).replace("\r\n", "\n").split("\n")]
    return "\n\n".join(p for p in paragraphs if p)


def read_rows(path: Path) -> list[tuple[str, str]]:
    """Read `(topic, content)` pairs from the sheet, in sheet order.

    Locates the header row by its labels rather than assuming a fixed offset
    (the table starts at B4 today, with a title in B2 and blank spacer rows),
    and fails loudly if the expected columns are gone — per the project's
    validate-at-the-boundary convention.
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    header_row = col_topic = col_content = None
    for row in sheet.iter_rows():
        labels = {(c.value or "").strip() if isinstance(c.value, str) else "": c.column
                  for c in row}
        if HEADER_TOPIC in labels and HEADER_CONTENT in labels:
            header_row = row[0].row
            col_topic, col_content = labels[HEADER_TOPIC], labels[HEADER_CONTENT]
            break
    if header_row is None:
        raise SystemExit(
            f"{path.name}: no header row with columns {HEADER_TOPIC!r} and "
            f"{HEADER_CONTENT!r} — has the sheet layout changed?")

    rows = []
    for r in range(header_row + 1, sheet.max_row + 1):
        topic = _cell_text(sheet.cell(r, col_topic).value)
        content = _cell_text(sheet.cell(r, col_content).value)
        if not topic and not content:
            continue                                   # trailing/spacer blank row
        if not topic or not content:                   # half-filled row: say so
            log.warning("row %d is incomplete (topic=%r, content=%d chars) — skipped",
                        r, topic[:40], len(content))
            continue
        rows.append((topic, content))
    if not rows:
        raise SystemExit(f"{path.name}: header found but no data rows below it")
    return rows


def group_rows(rows: list[tuple[str, str]]) -> dict[str, list[tuple[str, str]]]:
    """Group `(topic, content)` by the part before the dash, keeping sheet order.

    'Einspeiseanlagen – Zählerstand' -> group 'Einspeiseanlagen', sub-topic
    'Zählerstand'. A topic without a dash becomes its own group with no
    sub-topic (its content then sits directly under the h1).
    """
    groups: dict[str, list[tuple[str, str]]] = {}
    for topic, content in rows:
        parts = _TOPIC_SPLIT.split(topic, maxsplit=1)
        group, sub = (parts[0], parts[1]) if len(parts) == 2 else (topic, "")
        groups.setdefault(group, []).append((sub, content))
    return groups


def render(group: str, items: list[tuple[str, str]]) -> str:
    """One group -> one knowledge-base markdown file."""
    out = [f"# {HIERARCHY} - {group}", ""]
    for sub, content in items:
        if sub:
            out += [f"## {sub}", ""]
        out += [content, ""]
    return re.sub(r"\n{3,}", "\n\n", strip_links("\n".join(out))).rstrip() + "\n"


def _clean_generated(prefix: str) -> None:
    """Delete this script's own previous output before regenerating.

    A renamed or dropped group produces a different filename, so the superseded
    `.md` would linger in `static/`, keep being uploaded every week, and never
    be pruned — remote pruning only removes filenames *no longer produced
    locally* (`uploader.prune_stale`), and a stale local file still counts as
    produced. Scoped to the prefix this script owns, so hand-written pages
    (`Kundenportal.md`) and the PDF converter's output are untouched.
    """
    for stale in sorted(STATIC_DIR.glob(f"{prefix}_*.md")):
        stale.unlink()
        log.info("removed previous output %s", stale.name)


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    if not WORKBOOK.exists():
        log.error("workbook not found: %s", WORKBOOK)
        return 1

    # Render everything before touching static/ — a failure half-way must not
    # leave the folder emptied, or the run would upload a short set and
    # `prune_stale` would delete the rest of these files from the KB.
    groups = group_rows(read_rows(WORKBOOK))
    prefix = ascii_name(HIERARCHY)
    rendered = [(group, items, f"{prefix}_{ascii_name(group)}.md", render(group, items))
                for group, items in groups.items()]

    STATIC_DIR.mkdir(exist_ok=True)
    _clean_generated(prefix)
    for group, items, name, md in rendered:
        out = STATIC_DIR / name
        out.write_text(md, encoding="utf-8")
        log.info("%-22s %2d topic(s) -> %s (%d chars)", group, len(items), out.name, len(md))
        if len(md) > MAX_CHUNK:
            log.warning("%s is %d chars — over the %d-char cap, the API will split it",
                        out.name, len(md), MAX_CHUNK)
    log.info("%d groups, %d topics total", len(groups), sum(len(i) for i in groups.values()))
    return 0


if __name__ == "__main__":
    sys.exit(main())
